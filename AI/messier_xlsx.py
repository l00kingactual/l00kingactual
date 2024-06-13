from bs4 import BeautifulSoup
import requests
import openpyxl
import re
# import pymysql  # Uncomment when you're ready to use MySQL
import logging

# Initialize logging
logging.basicConfig(level=logging.INFO)

# Credentials object
class Credentials:
    def __init__(self, host, user, password, database):
        self.host = host
        self.user = user
        self.password = password
        self.database = database

# MySQL credentials
mysql_credentials = Credentials(host='localhost', user='username', password='password', database='your_database')

try:
    # Step 1: Scrape data from a webpage using BeautifulSoup
    url = "https://en.wikipedia.org/wiki/List_of_Messier_objects"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    table = soup.find('table', {'class': 'wikitable'})

    # Step 2: Save data to an Excel file using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in table.find_all('tr'):
        row_data = [re.sub(r'\[\d+\]', '', cell.text.strip()) for cell in row.find_all(['td', 'th'])]
        
        # Extract image URL if available
        img_tag = row.find('img')
        img_url = "https:" + img_tag['src'] if img_tag else "No Image"
        
        # Add image URL to row data
        row_data.append(img_url)
        
        ws.append(row_data)

    wb.save('messier_objects_with_images.xlsx')

    '''
    # Step 3: Insert data into MySQL database using pymysql
    connection = pymysql.connect(host=mysql_credentials.host, user=mysql_credentials.user, password=mysql_credentials.password, database=mysql_credentials.database)
    cursor = connection.cursor()

    # Create table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS messier_objects (
        id INT PRIMARY KEY AUTO_INCREMENT,
        messier_object VARCHAR(255),
        description VARCHAR(255),
        ra_dec VARCHAR(255),
        distance VARCHAR(255),
        constellation VARCHAR(255),
        associated_objects VARCHAR(255),
        wikiMessierObjectImageURL VARCHAR(255)
    );
    """)

    # Insert data
    wb = openpyxl.load_workbook('messier_objects_with_images.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        cursor.execute("INSERT INTO messier_objects (messier_object, description, ra_dec, distance, constellation, associated_objects, wikiMessierObjectImageURL) VALUES (%s, %s, %s, %s, %s, %s, %s)", row)

    connection.commit()
    '''

except Exception as e:
    logging.error(f"An error occurred: {e}")

'''
finally:
    # Close MySQL connection if it was opened
    if 'cursor' in locals():
        cursor.close()
    if 'connection' in locals():
        connection.close()
'''
