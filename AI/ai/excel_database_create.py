import os
import csv
import pandas as pd
from openpyxl import Workbook
from jinja2 import Environment, FileSystemLoader

# Function to create HTML table from tables_list.csv
def create_html_table(csv_file):
    with open(csv_file, 'r') as file:
        reader = csv.DictReader(file)
        rows = [row for row in reader]

    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template('template.html')

    with open('tables_list.html', 'w') as file:
        file.write(template.render(rows=rows))

import os
import pandas as pd
from openpyxl import Workbook

# Function to create HTML table from CSV
def create_html_table(csv_file):
    df = pd.read_csv(csv_file)
    html = df.to_html(index=False)
    with open("tables_list.html", "w") as f:
        f.write('<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">\n<title>Tables List</title>\n')
        f.write('<link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">\n')
        f.write('</head>\n<body>\n<div class="container mt-5">\n')
        f.write(html)
        f.write('\n</div>\n</body>\n</html>')

# Function to create Excel workbook and CSV files from CSV files
def create_excel_and_csv(directory_path, csv_output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet

    if not os.path.exists(csv_output_path):
        os.makedirs(csv_output_path)

    for filename in os.listdir(directory_path):
        if filename.endswith('.csv'):
            sheet_name = filename.replace('_cleaned', '').replace('.csv', '')
            df = pd.read_csv(os.path.join(directory_path, filename))

            # Create a new worksheet with the sheet_name
            ws = workbook.create_sheet(title=sheet_name)

            # Write DataFrame to worksheet
            for r_idx, row in enumerate(df.values, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Add header
            for c_idx, col in enumerate(df.columns, 1):
                ws.cell(row=1, column=c_idx, value=col)

            # Save DataFrame as CSV in the csv_output_path
            df.to_csv(os.path.join(csv_output_path, f"{sheet_name}.csv"), index=False)

    workbook.save('playground_data.xlsx')

if __name__ == "__main__":
    csv_file = "C:\\Users\\andy\\OneDrive\\Documents\\AstroAI\\ouchAstronomy\\AI\\ai\\tables_list.csv"
    directory_path = "C:\\Users\\andy\\OneDrive\\Documents\\AstroAI\\ouchAstronomy\\AI\\ai\\playground\\playground_data\\clean\\csv"
    csv_output_path = "C:\\Users\\andy\\OneDrive\\Documents\\AstroAI\\ouchAstronomy\\AI\\ai\\csv_data"

    create_html_table(csv_file)
    create_excel_and_csv(directory_path, csv_output_path)
