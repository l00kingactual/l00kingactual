import os
import re
import openpyxl
import logging

# Initialize logging
logging.basicConfig(level=logging.INFO)

# Function to clean special characters and '[any num or char]' from a string
def clean_data(value):
    try:
        # Remove special characters
        cleaned_value = re.sub(r"[^\w\s]", "", value)
        
        # Remove '[any num or char]'
        cleaned_value = re.sub(r"\[[^\]]+\]", "", cleaned_value)
        
        return cleaned_value
    except Exception as e:
        logging.error(f"Failed to clean data. Reason: {e}")
        return value

# Load the existing workbook
try:
    wb = openpyxl.load_workbook("data\\cggc_nwas_database.xlsx")
except Exception as e:
    logging.error(f"Failed to load workbook. Reason: {e}")
    exit(1)

# Iterate through each sheet in the workbook
for sheetname in wb.sheetnames:
    try:
        ws = wb[sheetname]
        logging.info(f"Processing sheet: {sheetname}")
        
        # Iterate through each cell in the sheet
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Clean the cell value
                    cell.value = clean_data(cell.value)
        
        logging.info(f"Completed cleaning sheet: {sheetname}")
    except Exception as e:
        logging.error(f"An error occurred while cleaning sheet {sheetname}. Reason: {e}")

# Save the cleaned workbook
try:
    wb.save("data\\cggc_nwas_database_cleaned.xlsx")
    logging.info("Cleaned workbook has been saved.")
except Exception as e:
    logging.error(f"Failed to save cleaned workbook. Reason: {e}")
