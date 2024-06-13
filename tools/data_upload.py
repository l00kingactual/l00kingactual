import os
import csv
import json
import pymysql
from openpyxl import load_workbook

# Database credentials
db_config = {
    'host': '213.171.200.30',
    'user': 'OuchAstronomy',
    'password': '@00E54m1sf1t?',
    'database': 'ouchAstronomy'
}

# Connect to the database
try:
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    print("INFO: Successfully connected to the database.")
except Exception as e:
    print(f"ERROR: Could not connect to the database. Reason: {e}")
    exit(1)

# Function to determine field types
def determine_field_types(rows):
    field_types = {}
    for row in rows:
        for key, value in row.items():
            if key not in field_types:
                field_types[key] = 'INT'
            if not str(value).isnumeric():
                field_types[key] = 'VARCHAR(255)'
    return field_types

# Initialize metadata dictionary
metadata = {}

# Read directory structure
directory_path = "data\\"
try:
    for root, dirs, files in os.walk(directory_path):
        for file in files:
            if file.endswith(('.csv', '.txt', '.md')):
                # Drop existing table
                table_name = os.path.splitext(file)[0]
                cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
                
                # Read file and create table
                with open(os.path.join(root, file), 'r') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                    field_types = determine_field_types(rows)
                    
                    # Update metadata
                    metadata[table_name] = list(field_types.keys())
                    
                    # Create table
                    create_table_query = f"CREATE TABLE {table_name} ("
                    create_table_query += ", ".join([f"{key} {value}" for key, value in field_types.items()])
                    create_table_query += ")"
                    
                    print(f"DEBUG: SQL Query: {create_table_query}")  # Debugging line
                    
                    cursor.execute(create_table_query)
                    print(f"INFO: Created table {table_name}")
                    
            elif file.endswith('.xlsx'):
                # Read Excel file
                workbook = load_workbook(filename=os.path.join(root, file))
                for sheet_name in workbook.sheetnames:
                    # Drop existing table
                    cursor.execute(f"DROP TABLE IF EXISTS {sheet_name}")
                    
                    # Read sheet and create table
                    sheet = workbook[sheet_name]
                    rows = list(sheet.iter_rows(values_only=True))
                    headers = rows[0]
                    data_rows = rows[1:]
                    
                    # Determine field types (simplified to INT or VARCHAR for this example)
                    field_types = ['INT' if all(isinstance(cell, int) for cell in col) else 'VARCHAR(255)' for col in zip(*data_rows)]
                    
                    # Update metadata
                    metadata[sheet_name] = headers
                    
                    # Create table
                    create_table_query = f"CREATE TABLE {sheet_name} ("
                    create_table_query += ", ".join([f"{headers[i]} {field_types[i]}" for i in range(len(headers))])
                    create_table_query += ")"
                    cursor.execute(create_table_query)
                    print(f"INFO: Created table {sheet_name}")

    # Write metadata to JSON file
    with open("metadata.json", 'w') as f:
        json.dump(metadata, f)
    print("INFO: Metadata saved to metadata.json")

except Exception as e:
    print(f"ERROR: An unexpected error occurred. Reason: {e}")

finally:
    # Close database connection
    cursor.close()
    connection.close()
    print("INFO: Database connection closed.")
