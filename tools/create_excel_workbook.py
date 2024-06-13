import os
import csv
import openpyxl
import logging
import warnings

# Initialize logging
logging.basicConfig(level=logging.INFO)

# Suppress warnings
warnings.filterwarnings("ignore")

# Initialize Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
wb.remove(ws)

# Define the root directory
root_directory = "data\\"

# Function to add CSV to Excel
def add_csv_to_excel(ws, csv_file_path):
    try:
        with open(csv_file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for r_idx, row in enumerate(reader, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value
        logging.info(f"Added CSV: {csv_file_path}")
    except Exception as e:
        logging.error(f"Failed to add CSV: {csv_file_path}. Reason: {e}")

# Iterate through each subdirectory in the root directory
for subdir in ['csv', 'md', 'txt', 'xlsx']:
    directory_path = os.path.join(root_directory, subdir)
    
    for filename in os.listdir(directory_path):
        try:
            file_path = os.path.join(directory_path, filename)
            sheet_title = os.path.splitext(filename)[0]
            sheet_title = (subdir + "_" + sheet_title)[:31]  # Truncate to 31 characters
            ws = wb.create_sheet(title=sheet_title)
            
            if filename.endswith('.csv'):
                add_csv_to_excel(ws, file_path)
            elif filename.endswith('.md') or filename.endswith('.txt'):
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    for r_idx, line in enumerate(f.readlines(), 1):
                        ws.cell(row=r_idx, column=1).value = line.strip()
                logging.info(f"Added text file: {file_path}")
            elif filename.endswith('.xlsx'):
                source_wb = openpyxl.load_workbook(file_path)
                for sheetname in source_wb.sheetnames:
                    source_ws = source_wb[sheetname]
                    target_ws = wb.create_sheet(title=(sheet_title + "_" + sheetname)[:31])
                    for row in source_ws.iter_rows():
                        for cell in row:
                            target_ws[cell.coordinate].value = cell.value
                logging.info(f"Added Excel workbook: {file_path}")
        except Exception as e:
            logging.error(f"An unexpected error occurred while processing {file_path}. Reason: {e}")

# Save the workbook
wb.save("data\\cggc_nwas_database.xlsx")
logging.info("Excel workbook has been created.")
